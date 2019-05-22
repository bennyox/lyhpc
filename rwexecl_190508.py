import openpyxl
from openpyxl import Workbook
import xlrd
import xlwt

import requests
from xlutils.copy import copy

filename = r'E:\丰年路仓库2019\2019-05-05\4月底盘库基础数据-20190505.xlsx'	#要读取的文件名
inwb = openpyxl.load_workbook(filename,data_only=True)  # 读文件

sheetnames = inwb.get_sheet_names()  # 获取读文件中所有的sheet，通过名字的方式
#print(sheetnames)
ws = inwb.get_sheet_by_name(sheetnames[0])  # 获取第一个sheet内容
#ws = inwb.get_sheet_by_name('配货清单')
print(ws)
# 获取sheet的最大行数和列数
rows = ws.max_row
cols = ws.max_column
print(rows,cols)
print(ws.cell(1,1).value)

outwb = xlrd.open_workbook("E:\\丰年路仓库2019\\2019-05-05\\表2：库存总金额统计表-2019-05-03———指导怎么补入库单--2019-05-08.xlsx")  # 打开一个将写的文件
#outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
#data = openpyxl.load_workbook("E:\\丰年路仓库2019\\2019-05-05\\表2：库存总金额统计表-2019-05-03———指导怎么补入库单--2019-05-08.xlsx")
copyfile=copy(outwb)
outws=outwb.sheet_by_name(u'库存总金额统计')
#outws=outwb.get_sheet_by_name(stname[0])                                      #下面有解释
rows2 = outws.nrows                  #使用wlrd的方法获取已有的的行数
cols2 = outws.ncols


print(outws,rows2,cols2)

for r in range(2,rows2):
	for c in range(2,rows):
		if outws.cell(r,2).value==ws.cell(c,1).value:
			print("yes")
			#outws.cell(r,3).value=ws.cell(c,2).value
			outws.write(r,3,ws.cell(c,2).value)
			outws.cell(r,4).value=ws.cell(c,3).value
			outws.cell(r,5).value=ws.cell(c,4).value
			outws.cell(r,6).value=ws.cell(c,5).value
			outws.cell(r,7).value=ws.cell(c,6).value
			outws.cell(r,8).value=ws.cell(c,7).value

saveExcel = "E:\\丰年路仓库2019\\2019-05-05\\表2：库存总金额统计表-2019-05-03———指导怎么补入库单--2019-05-05.xlsx"		#要写入的文件名
outwb.save(saveExcel)  # 一定要记得保存
